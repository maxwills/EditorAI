"""Convert uploaded files to Anthropic API content blocks.

Files arrive as raw bytes (via multipart upload) — no base64 encoding on the client.
Type detection order:
  1. File extension (from filename)
  2. Magic bytes (fallback when no filename or unknown extension)
  3. UTF-8 decode attempt (anything that reads as text)
  4. Skip with a warning (unreadable binary)

Supported formats
-----------------
Native (passed as-is to Claude):
  - Images     : .png .jpg .jpeg .gif .webp  → image block (Claude Vision)
  - PDF        : .pdf                         → document block

Converted server-side to text blocks:
  - Excel      : .xlsx .xls  → CSV text via openpyxl
  - Text-based : .csv .txt .json .xml .dxf .obj .stl (ASCII) .ifc .step .stp …

Unknown binary files are logged and skipped.
"""

import base64
import csv
import io
import os
from dataclasses import dataclass
from typing import Any

from app.utils.logging_utils import log


# ── Internal type ──────────────────────────────────────────────────────────────

@dataclass
class RawAttachment:
    """A file received from the client, already read into memory."""
    filename: str | None
    data: bytes


# ── Type detection ─────────────────────────────────────────────────────────────

_EXT_TO_MIME: dict[str, str] = {
    ".png":  "image/png",
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif":  "image/gif",
    ".webp": "image/webp",
    ".pdf":  "application/pdf",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls":  "application/vnd.ms-excel",
    ".csv":  "text/csv",
    ".txt":  "text/plain",
    ".json": "application/json",
    ".xml":  "application/xml",
    ".dxf":  "image/vnd.dxf",
    ".obj":  "model/obj",
    ".stl":  "model/stl",
    ".ifc":  "application/x-step",
    ".step": "model/step",
    ".stp":  "model/step",
}

#: (magic_prefix, mime) pairs — checked in order when extension is unavailable.
_MAGIC: list[tuple[bytes, str]] = [
    (b"\x89PNG",           "image/png"),
    (b"\xff\xd8\xff",      "image/jpeg"),
    (b"GIF8",              "image/gif"),
    (b"RIFF",              "image/webp"),   # bytes 8-12 must also be b"WEBP"
    (b"%PDF",              "application/pdf"),
    (b"PK\x03\x04",        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    (b"\xd0\xcf\x11\xe0",  "application/vnd.ms-excel"),
]


def _detect_media_type(filename: str | None, raw: bytes) -> str:
    if filename:
        ext = os.path.splitext(filename)[1].lower()
        if ext in _EXT_TO_MIME:
            return _EXT_TO_MIME[ext]

    for magic, mime in _MAGIC:
        if raw.startswith(magic):
            if magic == b"RIFF" and raw[8:12] != b"WEBP":
                continue
            return mime

    return "application/octet-stream"


# ── Format groups ──────────────────────────────────────────────────────────────

_IMAGE_TYPES: frozenset[str] = frozenset({
    "image/png", "image/jpeg", "image/gif", "image/webp",
})

_PDF_TYPES: frozenset[str] = frozenset({
    "application/pdf",
})

_EXCEL_TYPES: frozenset[str] = frozenset({
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/x-xls",
})


# ── Converters ─────────────────────────────────────────────────────────────────

def _excel_to_text(raw: bytes) -> str:
    import openpyxl  # lazy import — only needed when an Excel file is attached

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
        parts.append(f"[Sheet: {sheet_name}]\n{buf.getvalue().rstrip()}")
    wb.close()
    return "\n\n".join(parts)


# ── Public API ─────────────────────────────────────────────────────────────────

def to_content_blocks(prompt: str, attachments: list[RawAttachment]) -> list[dict[str, Any]]:
    """Build an Anthropic messages[].content list from a prompt and uploaded files.

    Returns the prompt string directly (not a list) when there are no
    attachments — preserving the existing single-string wire format.
    """
    if not attachments:
        return prompt  # type: ignore[return-value]

    blocks: list[dict[str, Any]] = [{"type": "text", "text": prompt}]

    for att in attachments:
        label = att.filename or "(unnamed)"
        mt = _detect_media_type(att.filename, att.data)
        log.debug("[attachment] '%s' detected as %s", label, mt)

        if mt in _IMAGE_TYPES:
            blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": mt,
                    "data": base64.b64encode(att.data).decode(),
                },
            })

        elif mt in _PDF_TYPES:
            blocks.append({
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": base64.b64encode(att.data).decode(),
                },
            })

        elif mt in _EXCEL_TYPES:
            try:
                text = _excel_to_text(att.data)
                blocks.append({"type": "text", "text": f"[File: {label}]\n{text}"})
            except Exception as exc:
                log.warning("[attachment] Excel parse failed for '%s': %s", label, exc)

        else:
            try:
                text = att.data.decode("utf-8")
                blocks.append({"type": "text", "text": f"[File: {label}]\n{text}"})
            except UnicodeDecodeError:
                log.warning(
                    "[attachment] '%s' (%s) is an unsupported binary format — skipped.", label, mt
                )

    return blocks
